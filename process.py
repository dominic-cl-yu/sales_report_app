
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_BASIS_EX_FTY = "ex_fty"
DATE_BASIS_CUSTOMER = "customer"

DATE_BASIS_COLUMN_MAP: Mapping[str, str] = {
    DATE_BASIS_EX_FTY: "Ex-Fty (Request Garment Delivery)",
    DATE_BASIS_CUSTOMER: "Customer Delivery Date",
}

SUMMARY_ORDER_TYPES: List[str] = ["SO", "AO", "Forecast-FR"]
METRICS: List[str] = ["Order Qty", "SAH", "Sales (USD)"]
FACTORY_ORDER: List[str] = ["CMBD", "CMSL", "CMVN"]

TEAM_SECTION_ORDER: List[str] = [
    "Sports",
    "Fancy",
    "SW",
    "Brands-COS",
    "Cotton Panty",
    "legging Reservation",
]
PRODUCT_SECTION_ORDER: List[str] = ["BOTTOM", "BRA", "OTHERS", "PANTIES", "TOP"]


class ReportError(Exception):
    """User-facing processing error."""


@dataclass(frozen=True)
class ReportConfig:
    report_date: Optional[str] = None
    date_basis: str = DATE_BASIS_EX_FTY


def _normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_key(value: object) -> str:
    text = _normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _clean_numeric(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce").fillna(0)
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "NaT": pd.NA})
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def _header_score(row: pd.Series) -> int:
    keys = {_normalize_key(value) for value in row.tolist() if _normalize_text(value)}
    expected = {
        "factory",
        "team",
        "producttype",
        "ordertype",
        "orderqty",
        "sah",
        "salesusd",
    }
    return len(keys & expected)


def _detect_header_row(excel_bytes: bytes, sheet_name: str, scan_rows: int = 80) -> int:
    preview = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=None, nrows=scan_rows)
    scores = {idx: _header_score(preview.iloc[idx]) for idx in range(len(preview))}
    best_idx = max(scores, key=scores.get)
    if scores[best_idx] < 4:
        raise ReportError(
            f"无法识别工作表 {sheet_name!r} 的表头行。前 {scan_rows} 行最高匹配分数只有 {scores[best_idx]}。"
        )
    return best_idx


def _read_sheet(excel_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    header_row = _detect_header_row(excel_bytes, sheet_name)
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=header_row)
    df = df.dropna(axis=1, how="all")
    return df


def _find_column(df: pd.DataFrame, aliases: Sequence[str], *, required: bool = True) -> Optional[str]:
    lookup = {_normalize_key(col): col for col in df.columns}
    for alias in aliases:
        hit = lookup.get(_normalize_key(alias))
        if hit is not None:
            return hit
    if required:
        raise ReportError(
            f"None of the expected columns were found: {list(aliases)}. "
            f"Available columns: {list(df.columns)}"
        )
    return None


def _pick_best_aofr_order_type_col(df: pd.DataFrame) -> str:
    candidates = [col for col in df.columns if _normalize_key(col).startswith("ordertype")]
    if not candidates:
        raise ReportError("AOFR 工作表中找不到 Order Type 列。")
    best_col = candidates[0]
    best_score = -1
    for col in candidates:
        values = df[col].dropna().astype(str).head(1000)
        score = sum(
            1
            for value in values
            if any(token in _normalize_text(value).lower() for token in ("ao", "forecast", "fr", "close"))
        )
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def _parse_month_label(value: object) -> Optional[str]:
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.strftime("%b-%y")

    if isinstance(value, (int, float)) and not pd.isna(value):
        if 20000 <= float(value) <= 80000:
            ts = pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")
            return ts.strftime("%b-%y")

    text = _normalize_text(value)
    if not text:
        return None

    ts = pd.to_datetime(text, errors="coerce")
    if pd.notna(ts):
        return ts.strftime("%b-%y")

    compact = re.sub(r"\s+", "", text)
    for fmt in ("%Y-%m", "%Y/%m", "%Y%m"):
        try:
            ts = pd.to_datetime(compact, format=fmt)
            return ts.strftime("%b-%y")
        except Exception:
            pass
    return None


def _month_from_year_month(year_value: object, month_value: object) -> Optional[str]:
    year = pd.to_numeric(pd.Series([year_value]), errors="coerce").iloc[0]
    month = pd.to_numeric(pd.Series([month_value]), errors="coerce").iloc[0]
    if pd.isna(year) or pd.isna(month):
        return None
    year = int(year)
    month = int(month)
    if not (1 <= month <= 12):
        return None
    return pd.Timestamp(year=year, month=month, day=1).strftime("%b-%y")


def _month_sort_key(label: str) -> Tuple[int, int]:
    ts = pd.to_datetime(label, format="%b-%y", errors="coerce")
    if pd.notna(ts):
        return (ts.year, ts.month)
    return (9999, 99)


def _continuous_month_columns(labels: Sequence[str]) -> List[str]:
    parsed = [pd.to_datetime(label, format="%b-%y", errors="coerce") for label in labels if label]
    parsed = [ts for ts in parsed if pd.notna(ts)]
    if not parsed:
        return []
    start = min(parsed).replace(day=1)
    end = max(parsed).replace(day=1)
    months: List[str] = []
    cursor = start
    while cursor <= end:
        months.append(cursor.strftime("%b-%y"))
        cursor = cursor + pd.offsets.MonthBegin(1)
    return months


def _normalize_team(value: object) -> str:
    text = _normalize_text(value)
    lowered = text.lower()
    if not lowered:
        return "ALL"
    if "legging reservation" in lowered:
        return "legging Reservation"
    if "fancy" in lowered:
        return "Fancy"
    if "sports-legging" in lowered or lowered == "sports":
        return "Sports"
    if "sports" in lowered and "legging" not in lowered:
        return "Sports"
    if lowered.startswith("sw") or lowered == "sw" or "sw-" in lowered or "sw " in lowered:
        return "SW"
    if "brands" in lowered or "cos" in lowered:
        return "Brands-COS"
    if "cotton" in lowered and "panty" in lowered:
        return "Cotton Panty"
    return text


def _normalize_product_type(value: object) -> str:
    text = _normalize_text(value)
    if not text:
        return "OTHERS"
    return text.upper()


def _classify_so_order_type(raw_value: object) -> str:
    lowered = _normalize_text(raw_value).lower()
    lowered = re.sub(r"[_-]+", " ", lowered)
    lowered = re.sub(r"\s+", " ", lowered).strip()
    if lowered in {"ao", "annual order", "ao close", "ao-close"} or lowered.startswith("ao "):
        return "AO"
    if "forecast" in lowered or lowered in {"fr", "forecast fr", "forecast-fr"}:
        return "Forecast-FR"
    # Meeting rule: Normal order / Speed order and other SO-like subtypes belong to SO
    return "SO"


def _classify_aofr_order_type(raw_value: object) -> str:
    lowered = _normalize_text(raw_value).lower()
    lowered = re.sub(r"[_-]+", " ", lowered)
    lowered = re.sub(r"\s+", " ", lowered).strip()
    if not lowered:
        return "AO"
    if "forecast" in lowered or lowered in {"fr", "forecast fr", "forecast-fr"}:
        return "Forecast-FR"
    return "AO"


def _derive_month(
    df: pd.DataFrame,
    *,
    date_basis: str,
    ex_fty_date_col: Optional[str],
    ex_fty_year_col: Optional[str],
    ex_fty_month_col: Optional[str],
    customer_date_col: Optional[str],
    customer_year_col: Optional[str],
    customer_month_col: Optional[str],
) -> pd.Series:
    labels: List[Optional[str]] = []
    for _, row in df.iterrows():
        if date_basis == DATE_BASIS_CUSTOMER:
            label = None
            if customer_date_col:
                label = _parse_month_label(row[customer_date_col])
            if label is None and customer_year_col and customer_month_col:
                label = _month_from_year_month(row[customer_year_col], row[customer_month_col])
        else:
            label = None
            if ex_fty_date_col:
                label = _parse_month_label(row[ex_fty_date_col])
            if label is None and ex_fty_year_col and ex_fty_month_col:
                label = _month_from_year_month(row[ex_fty_year_col], row[ex_fty_month_col])
        labels.append(label)
    return pd.Series(labels, index=df.index)


def _prepare_so_sheet(df: pd.DataFrame, date_basis: str) -> Tuple[pd.DataFrame, List[str]]:
    factory_col = _find_column(df, ["Factory"])
    team_col = _find_column(df, ["Team"], required=False)
    product_type_col = _find_column(df, ["Product Type"])
    order_type_col = _find_column(df, ["Order Type"])
    order_qty_col = _find_column(df, ["Order Qty"])
    sah_col = _find_column(df, ["SAH"])
    sales_col = _find_column(df, ["Sales (USD)", "Sales USD"])

    ex_fty_date_col = _find_column(
        df,
        ["Requested Garment Delivery (DeadLine ex-fty)", "Requested Garment Delivery", "Request Garment Delivery (DeadLine ex-fty)"],
        required=False,
    )
    ex_fty_year_col = _find_column(df, ["Ex-fty Year", "Ex Fty Year"], required=False)
    ex_fty_month_col = _find_column(df, ["Ex-fty Month", "Ex Fty Month"], required=False)
    customer_date_col = _find_column(df, ["Customer Delivery Date"], required=False)
    customer_year_col = _find_column(df, ["TOD Year", "TOD  Year"], required=False)
    customer_month_col = _find_column(df, ["TOD Month"], required=False)

    out = pd.DataFrame()
    out["Factory"] = df[factory_col].map(_normalize_text)
    out["Team"] = df[team_col].map(_normalize_team) if team_col else "ALL"
    out["Customer"] = "ALL"
    out["Product Type"] = df[product_type_col].map(_normalize_product_type)
    out["Raw Order Type"] = df[order_type_col].map(_normalize_text)
    out["Order Type"] = df[order_type_col].map(_classify_so_order_type)
    out["Month"] = _derive_month(
        df,
        date_basis=date_basis,
        ex_fty_date_col=ex_fty_date_col,
        ex_fty_year_col=ex_fty_year_col,
        ex_fty_month_col=ex_fty_month_col,
        customer_date_col=customer_date_col,
        customer_year_col=customer_year_col,
        customer_month_col=customer_month_col,
    )
    out["Order Qty"] = _clean_numeric(df[order_qty_col])
    out["SAH"] = _clean_numeric(df[sah_col])
    out["Sales (USD)"] = _clean_numeric(df[sales_col])
    out["Source Sheet"] = "SO"
    warnings: List[str] = []
    return out, warnings


def _prepare_aofr_sheet(df: pd.DataFrame, date_basis: str) -> Tuple[pd.DataFrame, List[str]]:
    factory_col = _find_column(df, ["Factory"])
    team_col = _find_column(df, ["Team"], required=False)
    product_type_col = _find_column(df, ["Product Type"])
    order_type_col = _pick_best_aofr_order_type_col(df)
    order_qty_col = _find_column(df, ["Order Qty"])
    sah_col = _find_column(df, ["SAH"])
    sales_col = _find_column(df, ["Sales (USD)", " Sales (USD)", "Sales USD"])

    ex_fty_date_col = _find_column(
        df,
        ["Request Garment Delivery (DeadLine ex-fty)", "Requested Garment Delivery (DeadLine ex-fty)", "Requested Garment Delivery"],
        required=False,
    )
    ex_fty_year_col = _find_column(df, ["Ex-fty Year", "Ex Fty Year"], required=False)
    ex_fty_month_col = _find_column(df, ["Ex-fty Month", "Ex Fty Month"], required=False)
    customer_date_col = _find_column(df, ["Customer Delivery Date"], required=False)
    customer_year_col = _find_column(df, ["TOD Year", "TOD  Year"], required=False)
    customer_month_col = _find_column(df, ["TOD Month"], required=False)

    out = pd.DataFrame()
    out["Factory"] = df[factory_col].map(_normalize_text)
    out["Team"] = df[team_col].map(_normalize_team) if team_col else "ALL"
    out["Customer"] = "ALL"
    out["Product Type"] = df[product_type_col].map(_normalize_product_type)
    out["Raw Order Type"] = df[order_type_col].map(_normalize_text)
    out["Order Type"] = df[order_type_col].map(_classify_aofr_order_type)
    out["Month"] = _derive_month(
        df,
        date_basis=date_basis,
        ex_fty_date_col=ex_fty_date_col,
        ex_fty_year_col=ex_fty_year_col,
        ex_fty_month_col=ex_fty_month_col,
        customer_date_col=customer_date_col,
        customer_year_col=customer_year_col,
        customer_month_col=customer_month_col,
    )
    out["Order Qty"] = _clean_numeric(df[order_qty_col])
    out["SAH"] = _clean_numeric(df[sah_col])
    out["Sales (USD)"] = _clean_numeric(df[sales_col])
    out["Source Sheet"] = "AOFR"
    warnings: List[str] = []
    return out, warnings


def _build_normalized_frame(excel_bytes: bytes, date_basis: str) -> Tuple[pd.DataFrame, Dict[str, object]]:
    warnings: List[str] = []

    so_df = _read_sheet(excel_bytes, "SO")
    aofr_df = _read_sheet(excel_bytes, "AOFR")

    so_norm, so_warnings = _prepare_so_sheet(so_df, date_basis)
    aofr_norm, aofr_warnings = _prepare_aofr_sheet(aofr_df, date_basis)
    warnings.extend(so_warnings)
    warnings.extend(aofr_warnings)

    frame = pd.concat([so_norm, aofr_norm], ignore_index=True)
    frame = frame[frame["Factory"].isin(FACTORY_ORDER)].copy()
    frame = frame[frame["Month"].notna()].copy()
    frame = frame[(frame["Order Qty"] != 0) | (frame["SAH"] != 0) | (frame["Sales (USD)"] != 0)].copy()

    if frame.empty:
        raise ReportError("没有可用于生成报表的数据。请检查工作簿内容和日期基准。")

    month_columns = _continuous_month_columns(sorted(frame["Month"].dropna().unique().tolist(), key=_month_sort_key))
    stats = {
        "rows_used": int(len(frame)),
        "factories": [fac for fac in FACTORY_ORDER if fac in frame["Factory"].unique().tolist()],
        "product_types": sorted(frame["Product Type"].dropna().unique().tolist()),
        "month_columns": month_columns,
        "warnings_list": warnings,
    }
    return frame, stats


def _report_date_label(report_date: Optional[str]) -> str:
    if report_date:
        ts = pd.to_datetime(report_date, errors="coerce")
        if pd.notna(ts):
            return ts.strftime("%b-%d")
        return str(report_date)
    return pd.Timestamp.today().strftime("%b-%d")


def _order_type_sort_key(value: str) -> int:
    try:
        return SUMMARY_ORDER_TYPES.index(value)
    except ValueError:
        return len(SUMMARY_ORDER_TYPES)


def _team_sort_key(value: str) -> Tuple[int, str]:
    try:
        return (TEAM_SECTION_ORDER.index(value), value)
    except ValueError:
        return (len(TEAM_SECTION_ORDER), value)


def _product_sort_key(value: str) -> Tuple[int, str]:
    try:
        return (PRODUCT_SECTION_ORDER.index(value), value)
    except ValueError:
        return (len(PRODUCT_SECTION_ORDER), value)


def _year_total_headers(month_columns: Sequence[str]) -> Tuple[List[str], Dict[int, List[str]]]:
    year_to_months: Dict[int, List[str]] = {}
    for month in month_columns:
        ts = pd.to_datetime(month, format="%b-%y", errors="coerce")
        if pd.isna(ts):
            continue
        year_to_months.setdefault(int(ts.year), []).append(month)
    year_total_headers = [f"{year} Ttl" for year in sorted(year_to_months)]
    return year_total_headers, year_to_months


def _append_year_totals(df: pd.DataFrame, month_columns: Sequence[str]) -> Tuple[pd.DataFrame, List[str]]:
    out = df.copy()
    year_total_headers, year_to_months = _year_total_headers(month_columns)
    for year in sorted(year_to_months):
        header = f"{year} Ttl"
        out[header] = out[year_to_months[year]].sum(axis=1)
    out["Ttl"] = out[year_total_headers].sum(axis=1) if year_total_headers else 0
    numeric_headers = [*month_columns, *year_total_headers, "Ttl"]
    return out, numeric_headers


def _write_row(ws, row_idx: int, values: Sequence[object], *, header: bool = False, header_fill: str = "D9EAF7") -> None:
    fill = PatternFill(fill_type="solid", fgColor=header_fill)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(vertical="center")
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if header:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = header_align
        else:
            cell.alignment = cell_align


def _write_dataframe_rows(ws, start_row: int, df: pd.DataFrame, *, header_fill: str = "D9EAF7") -> int:
    _write_row(ws, start_row, df.columns.tolist(), header=True, header_fill=header_fill)
    row = start_row + 1
    for _, item in df.iterrows():
        _write_row(ws, row, item.tolist(), header=False)
        row += 1
    return row - 1


def _write_subtotal_row(
    ws,
    row_idx: int,
    label_values: Sequence[object],
    *,
    first_numeric_col: int,
    last_numeric_col: int,
    data_start_row: int,
    data_end_row: int,
) -> None:
    _write_row(ws, row_idx, list(label_values), header=False)
    for col_idx in range(first_numeric_col, last_numeric_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=row_idx, column=col_idx, value=f"=SUBTOTAL(109,{col_letter}{data_start_row}:{col_letter}{data_end_row})").font = Font(bold=True)
    for col_idx in range(1, first_numeric_col):
        ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)


def _build_summary_df(frame: pd.DataFrame, factory: str, metric: str, month_columns: List[str], report_date: str) -> Tuple[pd.DataFrame, List[str]]:
    subset = frame[frame["Factory"] == factory]
    pivot = subset.pivot_table(index="Order Type", columns="Month", values=metric, aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(index=SUMMARY_ORDER_TYPES, fill_value=0)
    pivot = pivot.reindex(columns=month_columns, fill_value=0)
    pivot = pivot.reset_index()
    pivot.insert(0, "Factory", factory)
    pivot.insert(2, "Date", report_date)
    numeric = pivot[month_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
    pivot.loc[:, month_columns] = numeric
    pivot, numeric_headers = _append_year_totals(pivot, month_columns)
    ordered_cols = ["Factory", "Order Type", "Date", *numeric_headers]
    pivot = pivot.reindex(columns=ordered_cols)
    return pivot, numeric_headers


def _build_all_detail_df(frame: pd.DataFrame, factory: str, metric: str, month_columns: List[str], report_date: str) -> Tuple[pd.DataFrame, List[str]]:
    subset = frame[frame["Factory"] == factory].copy()
    detail = subset.pivot_table(
        index=["Factory", "Order Type", "Team", "Customer", "Product Type"],
        columns="Month",
        values=metric,
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    if detail.empty:
        detail = pd.DataFrame(columns=["Factory", "Order Type", "Team", "Customer", "Product Type", "Date", *month_columns])
    else:
        detail.insert(5, "Date", report_date)
        detail = detail.reindex(columns=["Factory", "Order Type", "Team", "Customer", "Product Type", "Date", *month_columns], fill_value=0)
        detail = detail.sort_values(
            by=["Team", "Product Type", "Order Type"],
            key=lambda s: s.map(dict((t, i) for i, t in enumerate(TEAM_SECTION_ORDER))) if s.name == "Team"
                else s.map(dict((p, i) for i, p in enumerate(PRODUCT_SECTION_ORDER))) if s.name == "Product Type"
                else s.map(_order_type_sort_key) if s.name == "Order Type"
                else s,
            kind="stable",
        ).reset_index(drop=True)
    detail, numeric_headers = _append_year_totals(detail, month_columns)
    ordered_cols = ["Factory", "Order Type", "Team", "Customer", "Product Type", "Date", *numeric_headers]
    detail = detail.reindex(columns=ordered_cols, fill_value=0)
    return detail, numeric_headers


def _build_section_df(frame: pd.DataFrame, factory: str, team: str, product_type: str, metric: str, month_columns: List[str], report_date: str) -> Tuple[pd.DataFrame, List[str]]:
    subset = frame[(frame["Factory"] == factory) & (frame["Team"] == team) & (frame["Product Type"] == product_type)].copy()
    pivot = subset.pivot_table(index="Order Type", columns="Month", values=metric, aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(columns=month_columns, fill_value=0)
    pivot = pivot.reset_index()
    pivot.insert(0, "Factory", factory)
    pivot.insert(2, "Team", team)
    pivot.insert(3, "Customer", "ALL")
    pivot.insert(4, "Product Type", product_type)
    pivot.insert(5, "Date", report_date)
    if not pivot.empty:
        pivot = pivot.sort_values(by="Order Type", key=lambda s: s.map(_order_type_sort_key), kind="stable").reset_index(drop=True)
    pivot, numeric_headers = _append_year_totals(pivot, month_columns)
    ordered_cols = ["Factory", "Order Type", "Team", "Customer", "Product Type", "Date", *numeric_headers]
    pivot = pivot.reindex(columns=ordered_cols, fill_value=0)
    return pivot, numeric_headers


def _section_candidates(frame: pd.DataFrame, factory: str) -> List[Tuple[str, str]]:
    subset = frame[frame["Factory"] == factory][["Team", "Product Type"]].drop_duplicates().copy()
    if subset.empty:
        return []
    subset["_team_order"] = subset["Team"].map(lambda x: _team_sort_key(x)[0])
    subset["_team_name"] = subset["Team"].map(lambda x: _team_sort_key(x)[1])
    subset["_product_order"] = subset["Product Type"].map(lambda x: _product_sort_key(x)[0])
    subset["_product_name"] = subset["Product Type"].map(lambda x: _product_sort_key(x)[1])
    subset = subset.sort_values(["_team_order", "_team_name", "_product_order", "_product_name"], kind="stable")
    return list(subset[["Team", "Product Type"]].itertuples(index=False, name=None))


def _auto_fit(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value) + 2, 40))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _render_workbook(frame: pd.DataFrame, report_date: str, month_columns: List[str]) -> bytes:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    factories = [fac for fac in FACTORY_ORDER if fac in frame["Factory"].unique().tolist()]
    for metric in METRICS:
        for factory in factories:
            ws = wb.create_sheet(f"{factory} - {metric}"[:31])

            # Summary block
            ws.cell(row=1, column=1, value=f"{factory} — ALL --- {metric} (All Teams + Product Types)").font = Font(bold=True)
            summary_df, summary_numeric_headers = _build_summary_df(frame, factory, metric, month_columns, report_date)
            summary_last_data_row = _write_dataframe_rows(ws, 2, summary_df)
            summary_total_row = summary_last_data_row + 1
            _write_subtotal_row(
                ws,
                summary_total_row,
                [factory, "Total", report_date],
                first_numeric_col=4,
                last_numeric_col=3 + len(summary_numeric_headers),
                data_start_row=3,
                data_end_row=summary_last_data_row,
            )

            # All-detail block
            detail_title_row = summary_total_row + 4
            ws.cell(row=detail_title_row, column=1, value=f"{factory} — ALL --- {metric} (All Teams + Product Types) - Detail").font = Font(bold=True)
            detail_df, detail_numeric_headers = _build_all_detail_df(frame, factory, metric, month_columns, report_date)
            detail_header_row = detail_title_row + 1
            detail_last_data_row = _write_dataframe_rows(ws, detail_header_row, detail_df)
            detail_total_row = detail_last_data_row + 1
            if detail_df.empty:
                _write_row(ws, detail_total_row, [factory, "Total", "Total", "ALL", "ALL", report_date], header=False)
            else:
                _write_subtotal_row(
                    ws,
                    detail_total_row,
                    [factory, "Total", "Total", "ALL", "ALL", report_date],
                    first_numeric_col=7,
                    last_numeric_col=6 + len(detail_numeric_headers),
                    data_start_row=detail_header_row + 1,
                    data_end_row=detail_last_data_row,
                )

            # Team/Product sections
            next_title_row = detail_total_row + 4
            for team, product_type in _section_candidates(frame, factory):
                section_df, section_numeric_headers = _build_section_df(frame, factory, team, product_type, metric, month_columns, report_date)
                if section_df.empty:
                    continue
                ws.cell(row=next_title_row, column=1, value=f"{factory} — {team} --- {product_type} ({metric})").font = Font(bold=True)
                section_header_row = next_title_row + 1
                section_last_data_row = _write_dataframe_rows(ws, section_header_row, section_df)
                section_total_row = section_last_data_row + 1
                _write_subtotal_row(
                    ws,
                    section_total_row,
                    [factory, "Total", "Total", "ALL", product_type, report_date],
                    first_numeric_col=7,
                    last_numeric_col=6 + len(section_numeric_headers),
                    data_start_row=section_header_row + 1,
                    data_end_row=section_last_data_row,
                )
                next_title_row = section_total_row + 4

            ws.freeze_panes = "A2"
            _auto_fit(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pivot_report_from_upload(
    *,
    excel_bytes: bytes,
    filename: Optional[str] = None,
    report_date: Optional[str] = None,
    date_basis: str = DATE_BASIS_EX_FTY,
) -> Tuple[bytes, Dict[str, object]]:
    if not excel_bytes:
        raise ReportError("上传文件为空。")
    if date_basis not in DATE_BASIS_COLUMN_MAP:
        raise ReportError(f"不支持的日期基准：{date_basis!r}")

    frame, stats = _build_normalized_frame(excel_bytes, date_basis)
    report_date_label = _report_date_label(report_date)
    workbook_bytes = _render_workbook(frame, report_date_label, stats["month_columns"])

    warnings_text = ""
    if stats["warnings_list"]:
        warnings_text = "\n".join(dict.fromkeys(str(x) for x in stats["warnings_list"] if str(x).strip()))

    result_stats = {
        "rows_used": stats["rows_used"],
        "factories": stats["factories"],
        "product_types": stats["product_types"],
        "report_date": report_date_label,
        "date_column": DATE_BASIS_COLUMN_MAP[date_basis],
        "warnings": warnings_text,
        "filename": filename or "pivot_report.xlsx",
    }
    return workbook_bytes, result_stats
